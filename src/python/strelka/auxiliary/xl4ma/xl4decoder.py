# Authors: Ryan Borre

import tempfile
import xlrd2
from openpyxl.workbook import Workbook
from pyxlsb2 import open_workbook
from pyxlsb2.formula import Formula
from pyxlsb2.records import ErrorValue
from os import devnull
from strelka.auxiliary.xl4ma.xl4interpreter import Interpreter


def _sanitize_results(results):
    sanitized_results = list()

    for result in results or []:
        if isinstance(result, str):
            sanitized = ", ".join([param.lstrip(' ').rstrip(' ') for param in str(result).replace('"', '').split(',')])
            sanitized_results.append(sanitized)

    return sanitized_results


# XLS
def _decode_xls(file_path, defined_names):
    wb = xlrd2.open_workbook(file_path, logfile=open(devnull, 'w'))
    book = Workbook()
    for sheet_name in wb.sheet_names():
        sheet_xls = wb.sheet_by_name(sheet_name)
        book_sheet = book.create_sheet(sheet_name)

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                try:
                    if wb[sheet_name].cell(row, col).ctype in (3, 4, 5):
                        book_sheet.cell(row+1, col+1, f"={wb[sheet_name].cell(row, col).formula}")
                    elif wb[sheet_name].cell(row, col).ctype in (1, 2):
                        book_sheet.cell(row+1, col+1, wb[sheet_name].cell(row, col).value)
                except:
                    pass
    temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    book.save(temp_file.name)
    book.close()

    return _sanitize_results(Interpreter(defined_names).calculate(temp_file))


# XLSB
def _decode_xlsb(file_path, defined_names):
    wb = open_workbook(file_path)
    book = Workbook()
    for sheet in wb.sheets:
        book_sheet = book.create_sheet(sheet.name)
        try:
            for row in wb.get_sheet_by_name(sheet.name):
                for cell in row:
                    if isinstance(cell.value, ErrorValue):
                        formula = Formula.parse(cell.formula).stringify(wb)
                        if '(' in formula and ')' in formula:
                            book_sheet.cell(cell.row.num + 1, cell.col + 1, f"={formula}")
                        else:
                            book_sheet.cell(cell.row.num + 1, cell.col + 1, formula)
                    elif isinstance(cell.formula, bytes) and (isinstance(cell.value, bool)):
                        book_sheet.cell(cell.row.num + 1, cell.col + 1, f"={Formula.parse(cell.formula).stringify(wb)}")
                    elif cell.value:
                        if isinstance(cell.value, int):
                            book_sheet.cell(cell.row.num + 1, cell.col + 1, int(cell.value))
                        elif isinstance(cell.value, float):
                            book_sheet.cell(cell.row.num + 1, cell.col + 1, float(cell.value))
                        elif isinstance(cell.value, str):
                            book_sheet.cell(cell.row.num + 1, cell.col + 1, str(cell.value).rstrip('\x00'))
        except:
            pass
    temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    book.save(temp_file.name)
    book.close()

    return _sanitize_results(Interpreter(defined_names).calculate(temp_file))


# XLSM
def _decode_xlsm(file_path, defined_names):
    with tempfile.NamedTemporaryFile(suffix=f".xlsm", delete=False) as temp_file, open(file_path, 'rb') as fp:
        temp_file.write(fp.read())

    return _sanitize_results(Interpreter(defined_names).calculate(temp_file))


def decode(file_path, file_type, defined_names):
    if file_type == 'xls':
        return _decode_xls(file_path, defined_names)
    if file_type == 'xlsb':
        return _decode_xlsb(file_path, defined_names)
    if file_type == 'xlsm':
        return _decode_xlsm(file_path, defined_names)
