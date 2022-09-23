# Authors: Ryan Borre

import re
from openpyxl import load_workbook


class XLSMWrapper:

    def __init__(self, file_path):
        try:
            self.workbook = load_workbook(file_path, read_only=False, keep_vba=True)
        except Exception as e:
            return

    def get_defined_names(self):
        defined_names = list()
        for name_obj in self.workbook.defined_names.definedName:
            defined_names.append(name_obj.name)

        return defined_names

    def parse_sheets(self, file_path):
        results = {
            'sheets': []
        }
        formula_count = 0
        value_count = 0
        sheet_count = 0

        for sheet in self.workbook.sheetnames:
            sheet_count += 1
            formulas = []
            values = []
            for row in self.workbook[sheet].rows:
                for cell in row:
                    if cell.value:
                        if cell.data_type == "f":
                            formula_count += 1
                            formulas.append({
                                "cell": cell.coordinate,
                                "value": cell.value
                            })
                            if re.match("^=?\w+?\(\)$", cell.value):
                                worksheet = self.workbook.get_sheet_by_name(sheet)
                                worksheet[cell.coordinate] = ""
                        elif cell.data_type == "n":
                            value_count += 1
                            values.append({
                                "cell": cell.coordinate,
                                "value": cell.value
                            })
                        elif cell.data_type == "s":
                            value_count += 1
                            values.append({
                                "cell": cell.coordinate,
                                "value": cell.value
                            })

            results['sheets'].append({
                # 'sheet_number': sheet.sheetId,
                'sheet_name': sheet,
                # 'sheet_type': sheet.type.upper(),
                'visibility': self.workbook[sheet].sheet_state.upper(),
                'formulas': formulas,
                'values': values,
            })

        results['defined_names'] = self.get_defined_names()
        results['meta'] = {
            'formulas': formula_count,
            'values': value_count,
            'sheets': sheet_count,
            'defined_names': len(results['defined_names'])
        }

        self.workbook.save(file_path)

        return results
